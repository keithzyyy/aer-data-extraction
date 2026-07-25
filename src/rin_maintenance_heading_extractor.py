"""Heading-driven extraction helpers for AER RIN maintenance worksheets."""

from __future__ import annotations

import json
import re
import unicodedata
from dataclasses import dataclass
from pathlib import Path
from typing import Any

import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet


DEFAULT_SCHEMA_PATH = (
    Path(__file__).resolve().parents[1]
    / "config"
    / "rin_maintenance_expected_schema.json"
)
SUPPORTED_WORKBOOK_SUFFIXES = {".xlsx", ".xlsm"}
REQUIRED_SECTION_NAMES = {"descriptor_metrics", "cost_metrics"}


class MaintenanceExtractionError(RuntimeError):
    """Raised when required workbook structure cannot be resolved safely."""


@dataclass(frozen=True)
class CellLocation:
    """A zero-based grid location with its one-based Excel coordinate."""

    row: int
    column: int
    coordinate: str


@dataclass(frozen=True)
class CellRegion:
    """Inclusive zero-based bounds for a worksheet section."""

    min_row: int
    max_row: int
    min_column: int
    max_column: int


@dataclass(frozen=True)
class HeaderMatch:
    """A canonical field matched to a leaf heading and its heading path."""

    name: str
    location: CellLocation
    raw_text: str
    path_texts: tuple[str, ...]
    path_locations: tuple[CellLocation, ...]


@dataclass
class MaintenanceExtractionResult:
    """Canonical tables and structural details extracted from one workbook."""

    workbook_path: Path
    sheet_name: str
    reporting_period: str
    template_date: str | None
    layout_profile: str
    descriptor_metrics: pd.DataFrame
    cost_metrics: pd.DataFrame
    header_locations: dict[str, dict[str, str]]
    warnings: list[str]


def _excel_coordinate(row: int, column: int) -> str:
    """Convert a zero-based grid position to an Excel A1 coordinate."""
    return f"{get_column_letter(column + 1)}{row + 1}"


def _location(row: int, column: int) -> CellLocation:
    """Construct a cell location from zero-based coordinates."""
    return CellLocation(
        row=row,
        column=column,
        coordinate=_excel_coordinate(row, column),
    )


def _is_blank(value: object) -> bool:
    """Return whether a scalar cell value should be treated as blank."""
    if value is None:
        return True

    # Treat empty strings and pandas' scalar missing values as blank.
    if isinstance(value, str):
        return not value.strip()

    try:
        return bool(pd.isna(value))
    except (TypeError, ValueError):
        return False


def _display_text(value: object) -> str:
    """Collapse presentation whitespace while preserving submitted casing."""
    if _is_blank(value):
        return ""
    return re.sub(r"\s+", " ", str(value)).strip()


def normalize_heading(value: object) -> str:
    """Normalize heading text without using fuzzy or positional matching."""
    if _is_blank(value):
        return ""

    # Normalize Unicode punctuation that otherwise creates false deviations.
    text = unicodedata.normalize("NFKC", str(value))
    text = text.translate(
        str.maketrans(
            {
                "\u2018": "'",
                "\u2019": "'",
                "\u2013": "-",
                "\u2014": "-",
            }
        )
    )

    # Normalize presentation line breaks and inconsistent slash spacing.
    text = re.sub(r"\s*/\s*", " / ", text)
    text = re.sub(r"\s+", " ", text).strip()

    return text.upper()


def _validate_match_rule(rule: object, context: str) -> None:
    """Validate one exact- or prefix-match rule from the JSON schema."""
    if not isinstance(rule, dict):
        raise MaintenanceExtractionError(f"{context} must be a JSON object")

    match_mode = rule.get("match")
    aliases = rule.get("aliases")

    if match_mode not in {"exact", "prefix"}:
        raise MaintenanceExtractionError(
            f"{context}.match must be 'exact' or 'prefix'"
        )
    if not isinstance(aliases, list) or not aliases:
        raise MaintenanceExtractionError(
            f"{context}.aliases must be a non-empty list"
        )
    if not all(isinstance(alias, str) and alias.strip() for alias in aliases):
        raise MaintenanceExtractionError(
            f"{context}.aliases must contain non-empty strings"
        )


def _validate_section_schema(section_name: str, section: object) -> None:
    """Validate one configured maintenance section."""
    if not isinstance(section, dict):
        raise MaintenanceExtractionError(
            f"sections.{section_name} must be a JSON object"
        )

    # Validate the anchor expressions before any workbook is opened.
    anchor_patterns = section.get("anchor_patterns")
    if not isinstance(anchor_patterns, list) or not anchor_patterns:
        raise MaintenanceExtractionError(
            f"sections.{section_name}.anchor_patterns must be a non-empty list"
        )
    for pattern in anchor_patterns:
        if not isinstance(pattern, str) or not pattern:
            raise MaintenanceExtractionError(
                f"sections.{section_name}.anchor_patterns must contain strings"
            )
        try:
            re.compile(pattern)
        except re.error as error:
            raise MaintenanceExtractionError(
                f"Invalid anchor pattern for {section_name}: {pattern!r}"
            ) from error

    # Validate canonical fields and their semantic header paths.
    fields = section.get("fields")
    if not isinstance(fields, list) or not fields:
        raise MaintenanceExtractionError(
            f"sections.{section_name}.fields must be a non-empty list"
        )

    field_names = []
    for field_index, field in enumerate(fields):
        context = f"sections.{section_name}.fields[{field_index}]"
        if not isinstance(field, dict):
            raise MaintenanceExtractionError(f"{context} must be a JSON object")

        field_name = field.get("name")
        if not isinstance(field_name, str) or not field_name:
            raise MaintenanceExtractionError(
                f"{context}.name must be a non-empty string"
            )
        field_names.append(field_name)

        if not isinstance(field.get("required"), bool):
            raise MaintenanceExtractionError(
                f"{context}.required must be true or false"
            )

        header_path = field.get("header_path")
        if not isinstance(header_path, list) or not header_path:
            raise MaintenanceExtractionError(
                f"{context}.header_path must be a non-empty list"
            )
        for path_index, rule in enumerate(header_path):
            _validate_match_rule(rule, f"{context}.header_path[{path_index}]")

    if len(field_names) != len(set(field_names)):
        raise MaintenanceExtractionError(
            f"sections.{section_name}.fields contains duplicate names"
        )

    # Validate field references used to identify and check extracted rows.
    for key in ("identity_fields", "numeric_fields"):
        configured_names = section.get(key)
        if not isinstance(configured_names, list):
            raise MaintenanceExtractionError(
                f"sections.{section_name}.{key} must be a list"
            )
        unknown_names = set(configured_names) - set(field_names)
        if unknown_names:
            raise MaintenanceExtractionError(
                f"sections.{section_name}.{key} references unknown fields: "
                f"{sorted(unknown_names)}"
            )

    output_columns = section.get("output_columns")
    if not isinstance(output_columns, list) or not output_columns:
        raise MaintenanceExtractionError(
            f"sections.{section_name}.output_columns must be a non-empty list"
        )
    missing_field_outputs = set(field_names) - set(output_columns)
    if missing_field_outputs:
        raise MaintenanceExtractionError(
            f"sections.{section_name}.output_columns omits configured fields: "
            f"{sorted(missing_field_outputs)}"
        )


def load_expected_schema(schema_path: str | Path) -> dict[str, Any]:
    """Load and validate the semantic maintenance-heading schema."""
    resolved_path = Path(schema_path)

    # Fail with a project-specific message when the schema cannot be read.
    try:
        with resolved_path.open(encoding="utf-8") as schema_file:
            schema = json.load(schema_file)
    except (OSError, json.JSONDecodeError) as error:
        raise MaintenanceExtractionError(
            f"Could not load maintenance schema {resolved_path}: {error}"
        ) from error

    if not isinstance(schema, dict):
        raise MaintenanceExtractionError("Maintenance schema must be a JSON object")
    if schema.get("schema_version") != 1:
        raise MaintenanceExtractionError(
            "Maintenance schema_version must be 1"
        )

    # Validate sheet aliases and the regular expressions used for discovery.
    sheet_schema = schema.get("sheet")
    if not isinstance(sheet_schema, dict):
        raise MaintenanceExtractionError("Maintenance schema requires sheet settings")
    sheet_aliases = sheet_schema.get("aliases")
    if not isinstance(sheet_aliases, list) or not sheet_aliases:
        raise MaintenanceExtractionError("sheet.aliases must be a non-empty list")

    patterns = schema.get("patterns")
    if not isinstance(patterns, dict):
        raise MaintenanceExtractionError("Maintenance schema requires patterns")
    for pattern_name in ("reporting_period", "template_date"):
        pattern = patterns.get(pattern_name)
        if not isinstance(pattern, str) or not pattern:
            raise MaintenanceExtractionError(
                f"patterns.{pattern_name} must be a non-empty string"
            )
        try:
            re.compile(pattern)
        except re.error as error:
            raise MaintenanceExtractionError(
                f"Invalid patterns.{pattern_name}: {pattern!r}"
            ) from error

    # Require both independent tables and validate their field contracts.
    sections = schema.get("sections")
    if not isinstance(sections, dict):
        raise MaintenanceExtractionError("Maintenance schema requires sections")
    missing_sections = REQUIRED_SECTION_NAMES - set(sections)
    if missing_sections:
        raise MaintenanceExtractionError(
            f"Maintenance schema is missing sections: {sorted(missing_sections)}"
        )
    for section_name in REQUIRED_SECTION_NAMES:
        _validate_section_schema(section_name, sections[section_name])

    recognized_units = schema.get("recognized_source_units")
    if not isinstance(recognized_units, list):
        raise MaintenanceExtractionError(
            "recognized_source_units must be a list"
        )

    return schema


def _resolve_sheet_name(
    sheet_names: list[str],
    schema: dict[str, Any],
) -> str:
    """Resolve exactly one maintenance sheet using normalized aliases."""
    sheet_schema = schema["sheet"]
    aliases = {
        normalize_heading(sheet_schema.get("canonical_name", "")),
        *(normalize_heading(alias) for alias in sheet_schema["aliases"]),
    }
    aliases.discard("")

    matched_names = [
        sheet_name
        for sheet_name in sheet_names
        if normalize_heading(sheet_name) in aliases
    ]

    if len(matched_names) != 1:
        raise MaintenanceExtractionError(
            "Expected exactly one maintenance sheet matching "
            f"{sorted(aliases)}, found {matched_names}"
        )

    return matched_names[0]


def load_maintenance_sheet(
    workbook_path: Path,
    schema: dict[str, Any],
) -> tuple[pd.DataFrame, Worksheet, str]:
    """Load a bounded value grid and worksheet metadata without saving."""
    if not workbook_path.is_file():
        raise MaintenanceExtractionError(
            f"Workbook path is not a file: {workbook_path}"
        )
    if workbook_path.suffix.lower() not in SUPPORTED_WORKBOOK_SUFFIXES:
        raise MaintenanceExtractionError(
            f"Unsupported workbook extension: {workbook_path.suffix}"
        )

    # Use pandas to obtain a compact value grid with no assumed header row.
    try:
        with pd.ExcelFile(workbook_path, engine="openpyxl") as excel_file:
            sheet_name = _resolve_sheet_name(excel_file.sheet_names, schema)
            grid = pd.read_excel(
                excel_file,
                sheet_name=sheet_name,
                header=None,
                dtype=object,
            )
    except MaintenanceExtractionError:
        raise
    except Exception as error:
        raise MaintenanceExtractionError(
            f"Could not read workbook {workbook_path}: {error}"
        ) from error

    # Load the worksheet in normal in-memory mode so merged ranges are visible.
    # The workbook is never saved, and macros are never executed.
    try:
        workbook = load_workbook(
            workbook_path,
            read_only=False,
            data_only=False,
            keep_links=False,
        )
        worksheet = workbook[sheet_name]
    except Exception as error:
        raise MaintenanceExtractionError(
            f"Could not inspect workbook metadata for {workbook_path}: {error}"
        ) from error

    return grid, worksheet, sheet_name


def _iter_region_positions(region: CellRegion):
    """Yield each zero-based position inside an inclusive region."""
    for row in range(region.min_row, region.max_row + 1):
        for column in range(region.min_column, region.max_column + 1):
            yield row, column


def _raw_grid_value(grid: pd.DataFrame, row: int, column: int) -> object:
    """Read a bounded grid value without relying on worksheet dimensions."""
    if row < 0 or column < 0:
        return None
    if row >= grid.shape[0] or column >= grid.shape[1]:
        return None
    return grid.iat[row, column]


def _value_at(
    grid: pd.DataFrame,
    merged_values: dict[tuple[int, int], object],
    row: int,
    column: int,
) -> object:
    """Read a cell, falling back only to its actual merged-range value."""
    value = _raw_grid_value(grid, row, column)
    if not _is_blank(value):
        return value
    return merged_values.get((row, column))


def find_section_anchors(
    grid: pd.DataFrame,
    schema: dict[str, Any],
) -> dict[str, CellLocation]:
    """Locate exactly one configured anchor for each maintenance section."""
    anchors: dict[str, CellLocation] = {}

    # Search the bounded pandas grid rather than Excel's inflated used range.
    for section_name, section_schema in schema["sections"].items():
        matches = []
        patterns = [
            re.compile(pattern)
            for pattern in section_schema["anchor_patterns"]
        ]

        for row in range(grid.shape[0]):
            for column in range(grid.shape[1]):
                normalized = normalize_heading(grid.iat[row, column])
                if normalized and any(
                    pattern.search(normalized) for pattern in patterns
                ):
                    matches.append(_location(row, column))

        if len(matches) != 1:
            coordinates = [match.coordinate for match in matches]
            raise MaintenanceExtractionError(
                f"Expected one {section_name} anchor, found "
                f"{len(matches)} at {coordinates}"
            )

        anchors[section_name] = matches[0]

    return anchors


def _find_template_date(
    grid: pd.DataFrame,
    template_date_pattern: str,
) -> str | None:
    """Return the submitted template-date text when it is present."""
    pattern = re.compile(template_date_pattern)

    for row in range(grid.shape[0]):
        for column in range(grid.shape[1]):
            value = grid.iat[row, column]
            normalized = normalize_heading(value)
            if normalized and pattern.search(normalized):
                text = _display_text(value)
                parts = re.split(r"\s*-\s*", text, maxsplit=1)
                return parts[1].strip() if len(parts) == 2 else text

    return None


def detect_layout_profile(
    grid: pd.DataFrame,
    anchors: dict[str, CellLocation],
    schema: dict[str, Any],
) -> str:
    """Classify section orientation without using fixed cell coordinates."""
    descriptor_anchor = anchors["descriptor_metrics"]
    cost_anchor = anchors["cost_metrics"]

    # The legacy template places the cost table to the descriptor table's right.
    if (
        descriptor_anchor.row == cost_anchor.row
        and descriptor_anchor.column < cost_anchor.column
    ):
        return "legacy_side_by_side"

    # Stacked variants differ by a semantic template-date marker, not row number.
    if descriptor_anchor.row < cost_anchor.row:
        template_date = _find_template_date(
            grid,
            schema["patterns"]["template_date"],
        )
        return "stacked_revised" if template_date else "stacked_baseline"

    raise MaintenanceExtractionError(
        "Unsupported section orientation: "
        f"2.8.1 at {descriptor_anchor.coordinate}, "
        f"2.8.2 at {cost_anchor.coordinate}"
    )


def derive_section_regions(
    grid_shape: tuple[int, int],
    anchors: dict[str, CellLocation],
    layout_profile: str,
) -> dict[str, CellRegion]:
    """Derive rough section bounds from anchor orientation."""
    row_count, column_count = grid_shape
    if row_count == 0 or column_count == 0:
        raise MaintenanceExtractionError("Maintenance sheet is empty")

    descriptor_anchor = anchors["descriptor_metrics"]
    cost_anchor = anchors["cost_metrics"]

    # Stacked tables share columns but occupy separate vertical regions.
    if layout_profile in {"stacked_baseline", "stacked_revised"}:
        return {
            "descriptor_metrics": CellRegion(
                min_row=descriptor_anchor.row,
                max_row=cost_anchor.row - 1,
                min_column=0,
                max_column=column_count - 1,
            ),
            "cost_metrics": CellRegion(
                min_row=cost_anchor.row,
                max_row=row_count - 1,
                min_column=0,
                max_column=column_count - 1,
            ),
        }

    # The legacy tables share rows but occupy separate horizontal regions.
    if layout_profile == "legacy_side_by_side":
        return {
            "descriptor_metrics": CellRegion(
                min_row=descriptor_anchor.row,
                max_row=row_count - 1,
                min_column=descriptor_anchor.column,
                max_column=cost_anchor.column - 1,
            ),
            "cost_metrics": CellRegion(
                min_row=cost_anchor.row,
                max_row=row_count - 1,
                min_column=cost_anchor.column,
                max_column=column_count - 1,
            ),
        }

    raise MaintenanceExtractionError(
        f"Cannot derive regions for layout profile {layout_profile!r}"
    )


def build_merged_value_lookup(
    worksheet: Worksheet,
    grid_shape: tuple[int, int],
) -> dict[tuple[int, int], object]:
    """Map bounded merged cells to their top-left submitted values."""
    row_count, column_count = grid_shape
    merged_values: dict[tuple[int, int], object] = {}

    # Inspect only merge ranges that intersect the compact pandas value grid.
    for merged_range in worksheet.merged_cells.ranges:
        min_row = max(merged_range.min_row - 1, 0)
        max_row = min(merged_range.max_row - 1, row_count - 1)
        min_column = max(merged_range.min_col - 1, 0)
        max_column = min(merged_range.max_col - 1, column_count - 1)

        if min_row > max_row or min_column > max_column:
            continue

        top_left_value = worksheet.cell(
            merged_range.min_row,
            merged_range.min_col,
        ).value
        for row in range(min_row, max_row + 1):
            for column in range(min_column, max_column + 1):
                merged_values[(row, column)] = top_left_value

    return merged_values


def _matches_rule(value: object, rule: dict[str, Any]) -> bool:
    """Match one cell against an explicitly configured semantic rule."""
    normalized_value = normalize_heading(value)
    if not normalized_value:
        return False

    normalized_aliases = [
        normalize_heading(alias)
        for alias in rule["aliases"]
    ]
    if rule["match"] == "exact":
        return normalized_value in normalized_aliases
    return any(
        normalized_value.startswith(alias)
        for alias in normalized_aliases
    )


def _match_header_path(
    grid: pd.DataFrame,
    merged_values: dict[tuple[int, int], object],
    region: CellRegion,
    field_schema: dict[str, Any],
) -> list[HeaderMatch]:
    """Find leaf headings whose same-column ancestors match a header path."""
    header_path = field_schema["header_path"]
    leaf_rule = header_path[-1]
    candidates: list[HeaderMatch] = []

    for row, column in _iter_region_positions(region):
        # Match leaf headings only at their submitted top-left cells. Using a
        # merged fallback here would make one merged label look duplicated.
        leaf_value = _raw_grid_value(grid, row, column)
        if not _matches_rule(leaf_value, leaf_rule):
            continue

        # Walk upward in the same semantic column for multi-row parent headings.
        path_locations = [_location(row, column)]
        path_texts = [_display_text(leaf_value)]
        current_row = row
        path_is_valid = True

        for parent_rule in reversed(header_path[:-1]):
            parent_match = None
            for parent_row in range(current_row - 1, region.min_row - 1, -1):
                parent_value = _value_at(
                    grid,
                    merged_values,
                    parent_row,
                    column,
                )
                if _matches_rule(parent_value, parent_rule):
                    parent_match = (
                        _location(parent_row, column),
                        _display_text(parent_value),
                    )
                    break

            if parent_match is None:
                path_is_valid = False
                break

            parent_location, parent_text = parent_match
            path_locations.insert(0, parent_location)
            path_texts.insert(0, parent_text)
            current_row = parent_location.row

        if path_is_valid:
            candidates.append(
                HeaderMatch(
                    name=field_schema["name"],
                    location=_location(row, column),
                    raw_text=_display_text(leaf_value),
                    path_texts=tuple(path_texts),
                    path_locations=tuple(path_locations),
                )
            )

    return candidates


def _add_warning(warnings: list[str], message: str) -> None:
    """Append a warning once while preserving discovery order."""
    if message not in warnings:
        warnings.append(message)


def resolve_section_headers(
    grid: pd.DataFrame,
    merged_values: dict[tuple[int, int], object],
    region: CellRegion,
    section_schema: dict[str, Any],
    *,
    section_name: str = "section",
    warnings: list[str] | None = None,
) -> dict[str, HeaderMatch]:
    """Resolve canonical fields to semantic leaf headings."""
    collected_warnings = warnings if warnings is not None else []
    matches: dict[str, HeaderMatch] = {}

    for field_schema in section_schema["fields"]:
        field_name = field_schema["name"]
        candidates = _match_header_path(
            grid,
            merged_values,
            region,
            field_schema,
        )

        # Prefer the earliest matching header row over identically named data.
        if candidates:
            earliest_row = min(candidate.location.row for candidate in candidates)
            candidates = [
                candidate
                for candidate in candidates
                if candidate.location.row == earliest_row
            ]

        if not candidates:
            if field_schema["required"]:
                raise MaintenanceExtractionError(
                    f"{section_name}: required heading {field_name!r} "
                    "was not found"
                )
            _add_warning(
                collected_warnings,
                f"{section_name}: optional heading {field_name!r} "
                "was not found",
            )
            continue

        if len(candidates) > 1:
            coordinates = [
                candidate.location.coordinate
                for candidate in candidates
            ]
            raise MaintenanceExtractionError(
                f"{section_name}: heading {field_name!r} is ambiguous "
                f"at {coordinates}"
            )

        matches[field_name] = candidates[0]

    return matches


def _canonical_period(value: object, pattern: re.Pattern[str]) -> str | None:
    """Return a normalized reporting period when a cell matches the pattern."""
    normalized = normalize_heading(value)
    if not normalized or pattern.fullmatch(normalized) is None:
        return None
    return re.sub(r"\s*-\s*", "-", normalized)


def _find_section_periods(
    grid: pd.DataFrame,
    region: CellRegion,
    header_matches: dict[str, HeaderMatch],
    period_pattern: str,
) -> list[tuple[str, CellLocation]]:
    """Find period cells at or below the resolved section headers."""
    compiled_pattern = re.compile(period_pattern)
    minimum_row = max(
        match.location.row
        for match in header_matches.values()
    )
    periods = []

    for row in range(minimum_row, region.max_row + 1):
        for column in range(region.min_column, region.max_column + 1):
            period = _canonical_period(
                _raw_grid_value(grid, row, column),
                compiled_pattern,
            )
            if period is not None:
                periods.append((period, _location(row, column)))

    return periods


def resolve_reporting_period(
    grid: pd.DataFrame,
    section_regions: dict[str, CellRegion],
    header_matches: dict[str, dict[str, HeaderMatch]],
    period_pattern: str,
) -> str:
    """Require one consistent reporting period across both tables."""
    section_periods: dict[str, str] = {}

    for section_name, region in section_regions.items():
        periods = _find_section_periods(
            grid,
            region,
            header_matches[section_name],
            period_pattern,
        )
        if not periods:
            raise MaintenanceExtractionError(
                f"{section_name}: reporting period was not found"
            )

        unique_values = {period for period, _ in periods}
        unique_rows = {location.row for _, location in periods}
        if len(unique_values) != 1 or len(unique_rows) != 1:
            details = [
                f"{period} at {location.coordinate}"
                for period, location in periods
            ]
            raise MaintenanceExtractionError(
                f"{section_name}: reporting period is ambiguous: {details}"
            )

        section_periods[section_name] = next(iter(unique_values))

    if len(set(section_periods.values())) != 1:
        raise MaintenanceExtractionError(
            f"Section reporting periods disagree: {section_periods}"
        )

    return next(iter(section_periods.values()))


def _section_period_row(
    grid: pd.DataFrame,
    region: CellRegion,
    header_matches: dict[str, HeaderMatch],
    reporting_period: str,
) -> int:
    """Locate the single reporting-period row for one section."""
    matching_rows = set()
    minimum_row = max(
        match.location.row
        for match in header_matches.values()
    )

    for row in range(minimum_row, region.max_row + 1):
        for column in range(region.min_column, region.max_column + 1):
            normalized = normalize_heading(
                _raw_grid_value(grid, row, column)
            )
            if re.sub(r"\s*-\s*", "-", normalized) == reporting_period:
                matching_rows.add(row)

    if len(matching_rows) != 1:
        raise MaintenanceExtractionError(
            f"Could not resolve one reporting-period row for {reporting_period}"
        )
    return next(iter(matching_rows))


def _extract_currency_unit(header_match: HeaderMatch) -> str | None:
    """Extract the submitted currency unit from a direct-expenditure heading."""
    if not header_match.path_texts:
        return None

    parent_text = header_match.path_texts[0]
    parenthetical_unit = re.search(r"\(([^()]*)\)\s*$", parent_text)
    if parenthetical_unit:
        return parenthetical_unit.group(1).strip()
    if normalize_heading(parent_text).endswith("$"):
        return "$"
    return None


def _row_is_meaningful(
    grid: pd.DataFrame,
    row: int,
    values: dict[str, object],
    header_matches: dict[str, HeaderMatch],
    section_schema: dict[str, Any],
) -> bool:
    """Retain reported rows without keeping merged-label padding rows."""
    identity_fields = section_schema["identity_fields"]
    metric_fields = section_schema["numeric_fields"]

    # Any reported metric or secondary identity makes a row meaningful.
    if any(not _is_blank(values.get(field)) for field in metric_fields):
        return True
    if any(
        not _is_blank(values.get(field))
        for field in identity_fields[1:]
    ):
        return True

    # Preserve a heading-only activity at its submitted top-left cell, while
    # excluding blank rows that inherit only a merged activity label.
    primary_field = identity_fields[0]
    primary_match = header_matches.get(primary_field)
    if primary_match is None:
        return False
    raw_primary_value = _raw_grid_value(
        grid,
        row,
        primary_match.location.column,
    )
    return not _is_blank(raw_primary_value)


def extract_section_rows(
    grid: pd.DataFrame,
    worksheet: Worksheet,
    region: CellRegion,
    header_matches: dict[str, HeaderMatch],
    section_schema: dict[str, Any],
    reporting_period: str,
    workbook_path: Path,
    sheet_name: str,
    *,
    merged_values: dict[tuple[int, int], object] | None = None,
) -> pd.DataFrame:
    """Extract one canonical wide table from resolved semantic columns."""
    merged_lookup = (
        merged_values
        if merged_values is not None
        else build_merged_value_lookup(worksheet, grid.shape)
    )
    period_row = _section_period_row(
        grid,
        region,
        header_matches,
        reporting_period,
    )
    data_start_row = max(
        period_row,
        *(match.location.row for match in header_matches.values()),
    ) + 1

    # Derive section-level constants, such as the submitted expenditure unit.
    constant_values: dict[str, object] = {}
    currency_source_field = section_schema.get("currency_unit_source_field")
    if currency_source_field:
        currency_header = header_matches[currency_source_field]
        constant_values["source_currency_unit"] = _extract_currency_unit(
            currency_header
        )

    records = []
    field_schemas = {
        field["name"]: field
        for field in section_schema["fields"]
    }

    for row in range(data_start_row, region.max_row + 1):
        # Read only columns resolved from expected semantic headings.
        values = {}
        for field_name in field_schemas:
            header_match = header_matches.get(field_name)
            values[field_name] = (
                _value_at(
                    grid,
                    merged_lookup,
                    row,
                    header_match.location.column,
                )
                if header_match is not None
                else None
            )

        if not _row_is_meaningful(
            grid,
            row,
            values,
            header_matches,
            section_schema,
        ):
            continue

        # Attach section constants and source lineage for later concatenation.
        record = {
            "reporting_period": reporting_period,
            **values,
            **constant_values,
            "source_workbook": workbook_path.name,
            "source_sheet": sheet_name,
            "source_row": row + 1,
        }
        records.append(record)

    return pd.DataFrame(
        records,
        columns=section_schema["output_columns"],
    )


def _is_numeric_value(value: object) -> bool:
    """Check a nonblank scalar without coercing or discarding its source value."""
    if _is_blank(value):
        return True
    converted = pd.to_numeric(pd.Series([value]), errors="coerce").iloc[0]
    return not pd.isna(converted)


def validate_extracted_section(
    table: pd.DataFrame,
    section_schema: dict[str, Any],
    *,
    section_name: str = "section",
    recognized_source_units: list[str] | None = None,
) -> list[str]:
    """Validate canonical columns and return non-fatal value warnings."""
    expected_columns = section_schema["output_columns"]
    missing_columns = set(expected_columns) - set(table.columns)
    if missing_columns:
        raise MaintenanceExtractionError(
            f"{section_name}: extracted table is missing columns "
            f"{sorted(missing_columns)}"
        )
    if table.empty:
        raise MaintenanceExtractionError(
            f"{section_name}: no meaningful data rows were extracted"
        )

    # Every retained record must have at least one configured identity value.
    identity_fields = section_schema["identity_fields"]
    invalid_identity_rows = table[
        table[identity_fields].apply(
            lambda row: all(_is_blank(value) for value in row),
            axis=1,
        )
    ]
    if not invalid_identity_rows.empty:
        source_rows = invalid_identity_rows["source_row"].tolist()
        raise MaintenanceExtractionError(
            f"{section_name}: rows {source_rows} have no identity values"
        )

    warnings = []

    # Report, but preserve, nonnumeric source values in expected metric columns.
    for field_name in section_schema["numeric_fields"]:
        invalid_values = table.loc[
            ~table[field_name].map(_is_numeric_value),
            ["source_row", field_name],
        ]
        if not invalid_values.empty:
            details = [
                f"row {row.source_row}: {getattr(row, field_name)!r}"
                for row in invalid_values.itertuples(index=False)
            ]
            _add_warning(
                warnings,
                f"{section_name}: nonnumeric values in {field_name}: "
                + "; ".join(details),
            )

    # Recognize known source units without applying stage-two conversions.
    recognized_units = {
        normalize_heading(unit)
        for unit in (recognized_source_units or [])
    }
    for unit_field in section_schema.get("unit_fields", []):
        if unit_field not in table:
            continue
        submitted_units = {
            normalize_heading(value)
            for value in table[unit_field]
            if not _is_blank(value)
        }
        unknown_units = submitted_units - recognized_units
        if unknown_units:
            _add_warning(
                warnings,
                f"{section_name}: unrecognized source units in {unit_field}: "
                f"{sorted(unknown_units)}",
            )

    return warnings


def _unexplained_header_warnings(
    grid: pd.DataFrame,
    region: CellRegion,
    anchor: CellLocation,
    header_matches: dict[str, HeaderMatch],
    reporting_period: str,
    section_name: str,
) -> list[str]:
    """Report header-band text not consumed by anchors, paths, or periods."""
    period_row = _section_period_row(
        grid,
        region,
        header_matches,
        reporting_period,
    )
    consumed_positions = {(anchor.row, anchor.column)}
    for header_match in header_matches.values():
        consumed_positions.update(
            (location.row, location.column)
            for location in header_match.path_locations
        )

    unexplained = []
    for row in range(region.min_row, period_row + 1):
        for column in range(region.min_column, region.max_column + 1):
            value = _raw_grid_value(grid, row, column)
            normalized = normalize_heading(value)
            if not normalized or (row, column) in consumed_positions:
                continue
            if re.sub(r"\s*-\s*", "-", normalized) == reporting_period:
                continue
            unexplained.append(
                f"{_excel_coordinate(row, column)}={_display_text(value)!r}"
            )

    if not unexplained:
        return []
    return [
        f"{section_name}: unexplained header text: "
        + "; ".join(unexplained)
    ]


def extract_rin_maintenance(
    workbook_path: str | Path,
    *,
    schema_path: str | Path = DEFAULT_SCHEMA_PATH,
    print_warnings: bool = True,
) -> MaintenanceExtractionResult:
    """Extract canonical maintenance tables from one workbook path."""
    resolved_workbook_path = Path(workbook_path)
    schema = load_expected_schema(schema_path)
    print(f"[maintenance] Opening {resolved_workbook_path}")

    grid, worksheet, sheet_name = load_maintenance_sheet(
        resolved_workbook_path,
        schema,
    )
    workbook = worksheet.parent

    try:
        # Discover structure from semantic anchors rather than fixed positions.
        anchors = find_section_anchors(grid, schema)
        layout_profile = detect_layout_profile(grid, anchors, schema)
        regions = derive_section_regions(
            grid.shape,
            anchors,
            layout_profile,
        )
        merged_values = build_merged_value_lookup(
            worksheet,
            grid.shape,
        )

        # Resolve each section independently because their row sets can differ.
        warnings: list[str] = []
        section_headers = {}
        for section_name, region in regions.items():
            section_headers[section_name] = resolve_section_headers(
                grid,
                merged_values,
                region,
                schema["sections"][section_name],
                section_name=section_name,
                warnings=warnings,
            )

        reporting_period = resolve_reporting_period(
            grid,
            regions,
            section_headers,
            schema["patterns"]["reporting_period"],
        )

        # Extract canonical wide tables while preserving submitted labels/units.
        descriptor_metrics = extract_section_rows(
            grid,
            worksheet,
            regions["descriptor_metrics"],
            section_headers["descriptor_metrics"],
            schema["sections"]["descriptor_metrics"],
            reporting_period,
            resolved_workbook_path,
            sheet_name,
            merged_values=merged_values,
        )
        cost_metrics = extract_section_rows(
            grid,
            worksheet,
            regions["cost_metrics"],
            section_headers["cost_metrics"],
            schema["sections"]["cost_metrics"],
            reporting_period,
            resolved_workbook_path,
            sheet_name,
            merged_values=merged_values,
        )

        # Validate values without converting or suppressing submitted deviations.
        for warning in validate_extracted_section(
            descriptor_metrics,
            schema["sections"]["descriptor_metrics"],
            section_name="descriptor_metrics",
            recognized_source_units=schema["recognized_source_units"],
        ):
            _add_warning(warnings, warning)
        for warning in validate_extracted_section(
            cost_metrics,
            schema["sections"]["cost_metrics"],
            section_name="cost_metrics",
            recognized_source_units=schema["recognized_source_units"],
        ):
            _add_warning(warnings, warning)

        # Report header-band text that the expected schema did not explain.
        for section_name, region in regions.items():
            for warning in _unexplained_header_warnings(
                grid,
                region,
                anchors[section_name],
                section_headers[section_name],
                reporting_period,
                section_name,
            ):
                _add_warning(warnings, warning)

        # Preserve the exact leaf-heading coordinates used for extraction.
        header_locations = {
            section_name: {
                field_name: header_match.location.coordinate
                for field_name, header_match in matches.items()
            }
            for section_name, matches in section_headers.items()
        }
        currency_source_field = schema["sections"]["cost_metrics"].get(
            "currency_unit_source_field"
        )
        if currency_source_field:
            currency_match = section_headers["cost_metrics"][
                currency_source_field
            ]
            header_locations["cost_metrics"]["source_currency_unit"] = (
                currency_match.path_locations[0].coordinate
            )

        template_date = _find_template_date(
            grid,
            schema["patterns"]["template_date"],
        )
    finally:
        # Close file handles even when structural validation raises.
        workbook.close()

    print(
        f"[maintenance] Detected {layout_profile}; "
        f"descriptor rows={len(descriptor_metrics)}, "
        f"cost rows={len(cost_metrics)}"
    )
    if print_warnings:
        for warning in warnings:
            print(f"[maintenance] Warning: {warning}")

    return MaintenanceExtractionResult(
        workbook_path=resolved_workbook_path,
        sheet_name=sheet_name,
        reporting_period=reporting_period,
        template_date=template_date,
        layout_profile=layout_profile,
        descriptor_metrics=descriptor_metrics,
        cost_metrics=cost_metrics,
        header_locations=header_locations,
        warnings=warnings,
    )


__all__ = [
    "DEFAULT_SCHEMA_PATH",
    "MaintenanceExtractionError",
    "MaintenanceExtractionResult",
    "extract_rin_maintenance",
]
